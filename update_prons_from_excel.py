# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
update_prons_from_excel.py --- Betplay IGAC
Lee el Excel con pronósticos y los sube a Firebase con la perspectiva correcta.

CONVENCIÓN DE ALMACENAMIENTO EN FIREBASE (home_first perspective):
  - ida.home   = goles del equipo LOCAL en la IDA
  - ida.away   = goles del equipo VISITANTE en la IDA
  - vuelta.home = goles del equipo que fue LOCAL en la IDA (jugó de visitante en VUELTA)
  - vuelta.away = goles del equipo que fue VISITANTE en la IDA (jugó de local en VUELTA)

CONVERSIÓN DESDE EL EXCEL:
  - Fila IDA:    home = Goles Local     / away = Goles Visitante  (sin cambio)
  - Fila VUELTA: home = Goles Visitante / away = Goles Local      (INVERTIDO)
    → porque en la VUELTA, el away_first juega físicamente de local.
"""

import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
import json
import openpyxl
import os

SERVICE_ACCOUNT = "serviceAccountKey.json"
EXCEL_FILE      = "BETPLAY_UCL_Pronosticos_v3.xlsx"

# Mapeo: nombre del equipo LOCAL en la IDA del Excel → match_id en Firebase
IDA_LOCAL_TO_ID = {
    "París Saint-Germain": "par-che",
    "Galatasaray":          "gal-liv",
    "Real Madrid":          "rea-man",
    "Atalanta":             "ata-bay",
    "Newcastle United":     "new-bar",
    "Atlético de Madrid":   "atl-tot",
    "Bodø/Glimt":           "fk-spo",
    "Bayer Leverkusen":     "bay-ars",
}

def read_prons_from_excel(filepath: str) -> dict:
    """
    Devuelve un dict {participante: {match_id: {ida: {...}, vuelta: {...}}}}.
    Aplica la conversión de perspectiva para las filas VUELTA.
    """
    wb = openpyxl.load_workbook(filepath)
    all_prons = {}

    for participant in wb.sheetnames:
        ws = wb[participant]
        data: dict = {}
        current_id: str | None = None

        for row in ws.iter_rows(values_only=True):
            # Asegurar que la fila tenga al menos 5 columnas
            cells = (list(row) + [None] * 5)[:5]
            fase, local, visitante, goles_local, goles_visitante = cells

            if fase not in ("IDA", "VUELTA"):
                continue
            if goles_local is None or goles_visitante is None:
                continue

            if fase == "IDA":
                match_id = IDA_LOCAL_TO_ID.get(str(local).strip())
                if not match_id:
                    print(f"  [!] [{participant}] No se encontro match_id para IDA local='{local}'")
                    current_id = None
                    continue
                current_id = match_id
                data[match_id] = {
                    "ida":    {"home": int(goles_local),      "away": int(goles_visitante)},
                    "vuelta": {"home": None, "away": None},
                }

            elif fase == "VUELTA":
                if current_id is None:
                    continue
                # El equipo LOCAL físico en la VUELTA es el away_first del cruce.
                # Para mantener la perspectiva home_first:
                #   vuelta.home = goles del home_first (= Goles Visitante del Excel)
                #   vuelta.away = goles del away_first (= Goles Local del Excel)
                data[current_id]["vuelta"] = {
                    "home": int(goles_visitante),   # home_first fue visitante físico
                    "away": int(goles_local),        # away_first fue local físico
                }

        all_prons[participant] = data
        ida_ok    = sum(1 for v in data.values() if v["ida"]["home"] is not None)
        vuelta_ok = sum(1 for v in data.values() if v["vuelta"]["home"] is not None)
        print(f"  OK  {participant}: {ida_ok} IDAs / {vuelta_ok} VUELTAs leidos")

    return all_prons


def preview(all_prons: dict) -> None:
    """Imprime una tabla de verificación antes de subir."""
    print()
    print("  CRUCE        │ IDA  │ VTA  │ CRUCE        │ IDA  │ VTA  │ ...")
    print("  " + "─" * 70)
    match_ids = list(next(iter(all_prons.values())).keys())
    for mid in match_ids:
        parts = []
        for p in all_prons:
            d = all_prons[p].get(mid, {})
            ida = d.get("ida", {})
            vta = d.get("vuelta", {})
            i_s = f"{ida['home']}-{ida['away']}" if ida.get("home") is not None else "  ?"
            v_s = f"{vta['home']}-{vta['away']}" if vta.get("home") is not None else "  ?"
            parts.append(f"{p[:5]:<6}: {i_s:<5} {v_s:<5}")
        print(f"  {mid:<10} │ " + " │ ".join(parts))


def main():
    print(f"[INFO] Leyendo '{EXCEL_FILE}'...")
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] Archivo no encontrado: {EXCEL_FILE}")
        return

    all_prons = read_prons_from_excel(EXCEL_FILE)

    print("\n[PREVIEW] Pronosticos (home_first perspective):")
    preview(all_prons)

    print(f"\n[FIREBASE] Subiendo (fusion: mantiene cuartos/semis/final)...")
    if not firebase_admin._apps:
        cred = credentials.Certificate(SERVICE_ACCOUNT)
        firebase_admin.initialize_app(cred)
    db = firestore.client()

    for participant, oct_data in all_prons.items():
        doc_ref  = db.collection("betplay").document(f"pron_{participant}")
        existing = doc_ref.get()
        # Fusionar: conservar pronósticos de otras fases, sobreescribir solo los del Excel
        current = json.loads(existing.to_dict().get("data", "{}")) if existing.exists else {}
        current.update(oct_data)
        doc_ref.set({
            "data":       json.dumps(current),
            "updated_at": datetime.now().isoformat(),
        })
        print(f"  OK  pron_{participant} subido ({len(oct_data)} cruces)")

    print(f"\n[DONE] Listo - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
